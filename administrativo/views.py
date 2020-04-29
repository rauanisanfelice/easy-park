from django.shortcuts import render
from django.views.generic import View
from django.http import HttpResponse
from openpyxl import Workbook

from django.db.models.functions import TruncYear, TruncMonth
from django.db.models import Sum, Count

from usuario.models import Carteira, Parada, TipoNotificacao, Notificacao, Funcionario, Veiculo
from vendedor.models import PesquisaVeiculo, VendaFuncionario

import calendar
import datetime

HORAS_VALIDAS_DIA = 8
DIAS_UTEIS_SEMANA = 6

class Dashboard(View):
    template_name = 'dashboard-summary.html'
    
    def get(self, request):
        context = {}
        mes_dias = datetime.datetime.now().month
        ano_dias = datetime.datetime.now().year
        if 'limpar' in request.GET:
            ano = None
            mes = None
        else:    
            ano = request.GET.get('ano', None)
            mes = request.GET.get('mes', None)    
            if ano: ano_dias = int(ano)
            if mes: mes_dias = int(mes)
                

        # NOTIFICAÇÕES E INFRAÇÕES
        tp_notificacao = TipoNotificacao.objects.filter(id__in=[1,2])
        tp_infracao = TipoNotificacao.objects.filter(id__in=[3,4])

        # DESEMPENHO
        dias_mes = calendar.monthrange(ano_dias, mes_dias)[1]
        dias_validos = dias_mes - (DIAS_UTEIS_SEMANA % dias_mes)
        horas_validas = dias_validos * HORAS_VALIDAS_DIA

        # TODO DEIXAR UM FILTRO DE MES E ANO INICIALMENTE SELECIONADO PARA NAO FICAR PESADO AO CARREGAR A PAGINA
        # OBJETOS
        sum_valor_entradas = Carteira.objects.filter(tipo_lancamento='en')
        sum_valor_saidas = Parada.objects.all()
        sum_total_paradas = Parada.objects.all()
        sum_total_paradas_ativas = Parada.objects.filter(valido=True)
        total_veiculos_cadastrados = Veiculo.objects.filter(ativo=True)
        sum_total_notificacoes = Notificacao.objects.filter(tipo_notificacao__in=tp_notificacao)
        sum_total_notificacoes_lidas = Notificacao.objects.filter(tipo_notificacao__in=tp_notificacao, data_lida__isnull=False)
        sum_total_infracoes = Notificacao.objects.filter(tipo_notificacao__in=tp_infracao)
        sum_total_horas_paradas = Parada.objects.all()
        pesquisasFuncionarios = PesquisaVeiculo.objects.all()
        vendasFuncionarios = VendaFuncionario.objects.all()

        ##########################################
        # FILTROS
        if ano:
            # CREDITOS
            sum_valor_entradas = sum_valor_entradas.filter(data_insercao__year=ano)
            sum_valor_saidas = sum_valor_saidas.filter(data_parada__year=ano)

            # PARADAS
            sum_total_paradas = sum_total_paradas.filter(data_parada__year=ano)
            sum_total_paradas_ativas = sum_total_paradas_ativas.filter(data_parada__year=ano)

            # NOTIFICAÇÕES E INFRAÇÕES
            sum_total_notificacoes = sum_total_notificacoes.filter(data_notificacao__year=ano)
            sum_total_notificacoes_lidas = sum_total_notificacoes_lidas.filter(data_notificacao__year=ano)
            sum_total_infracoes = sum_total_infracoes.filter(data_notificacao__year=ano)

            # DESEMPENHO
            sum_total_horas_paradas = sum_total_horas_paradas.filter(data_parada__year=ano)

            # RANKING
            pesquisasFuncionarios = pesquisasFuncionarios.filter(data_pesquisa__year=ano)
            vendasFuncionarios = vendasFuncionarios.filter(data_venda__year=ano)
            
        if mes:
            # CREDITOS
            sum_valor_entradas = sum_valor_entradas.filter(data_insercao__month=mes)
            sum_valor_saidas = sum_valor_saidas.filter(data_parada__month=mes)

            # PARADAS
            sum_total_paradas = sum_total_paradas.filter(data_parada__month=mes)
            sum_total_paradas_ativas = sum_total_paradas_ativas.filter(data_parada__month=mes)

            # NOTIFICAÇÕES E INFRAÇÕES
            sum_total_notificacoes = sum_total_notificacoes.filter(data_notificacao__month=mes)
            sum_total_notificacoes_lidas = sum_total_notificacoes_lidas.filter(data_notificacao__month=mes)
            sum_total_infracoes = sum_total_infracoes.filter(data_notificacao__month=mes)

            # DESEMPENHO
            sum_total_horas_paradas = sum_total_horas_paradas.filter(data_parada__month=mes)

            # RANKING
            pesquisasFuncionarios = pesquisasFuncionarios.filter(data_pesquisa__month=mes)
            vendasFuncionarios = vendasFuncionarios.filter(data_venda__month=mes)

        
        ##########################################
        # AGGREGATE
        sum_valor_entradas = sum_valor_entradas.aggregate(Sum('valor'))['valor__sum']
        sum_valor_saidas = sum_valor_saidas.aggregate(Sum('quantidade_horas__valor'))['quantidade_horas__valor__sum']
        sum_total_paradas = sum_total_paradas.aggregate(Count('valido'))['valido__count']
        sum_total_paradas_ativas  = sum_total_paradas_ativas.aggregate(Count('valido'))['valido__count']
        total_veiculos_cadastrados = total_veiculos_cadastrados.distinct('placa').count()
        sum_total_notificacoes = sum_total_notificacoes.aggregate(Count('tipo_notificacao'))['tipo_notificacao__count']
        sum_total_notificacoes_lidas = sum_total_notificacoes_lidas.aggregate(Count('tipo_notificacao'))['tipo_notificacao__count']
        sum_total_infracoes = sum_total_infracoes.aggregate(Count('tipo_notificacao'))['tipo_notificacao__count']
        sum_total_horas_paradas = sum_total_horas_paradas.aggregate(Sum('quantidade_horas__horas'), Sum('quantidade_horas__minutos'))
        pesquisasFuncionarios = pesquisasFuncionarios.values('funcionario__first_name').annotate(total=Count('funcionario__first_name')).order_by('total')[:5]
        vendasFuncionarios = vendasFuncionarios.values('funcionario__first_name', 'parada__quantidade_horas__valor').annotate(total=Count('funcionario__first_name'), sum=Count('parada__quantidade_horas__valor')).order_by('funcionario__first_name')[:5]

        ##########################################
        # CALCULOS
        if sum_valor_entradas == None: sum_valor_entradas = 0
        if sum_valor_saidas == None: sum_valor_saidas = 0
        if sum_total_paradas == None: sum_total_paradas = 0
        if sum_total_paradas_ativas == None: sum_total_paradas_ativas = 0
        if sum_total_notificacoes == None: sum_total_notificacoes = 0
        if sum_total_notificacoes_lidas == None: sum_total_notificacoes_lidas = 0
        if sum_total_infracoes == None: sum_total_infracoes = 0

        # CREDITOS
        sum_valor_carteira = sum_valor_entradas - sum_valor_saidas

        # NOTIFICAÇÕES E INFRAÇÕES
        if sum_total_paradas ==  0:
            taxa_notificacoes = 0
            taxa_notificacoes_lidas = 0
            taxa_infracoes = 0
        else:
            taxa_notificacoes = float(((sum_total_notificacoes * 1) / sum_total_paradas) * 100)
            taxa_notificacoes_lidas = float(((sum_total_notificacoes_lidas * 1) / sum_total_notificacoes) * 100)
            taxa_infracoes = float(((sum_total_infracoes * 1) / sum_total_paradas) * 100)

        # DESEMPENHO
        if not sum_total_horas_paradas['quantidade_horas__horas__sum'] and not sum_total_horas_paradas['quantidade_horas__minutos__sum']:
            total_horas_paradas = 0
        else:
            total_horas_paradas = sum_total_horas_paradas['quantidade_horas__horas__sum'] + (sum_total_horas_paradas['quantidade_horas__minutos__sum'] / 60)
        delta_horas_paradas = total_horas_paradas - horas_validas

        ##########################################
        # RETORNO
        context["sum_valor_entradas"] = sum_valor_entradas
        context["sum_valor_carteira"] = sum_valor_carteira
        context["sum_valor_saidas"] = sum_valor_saidas
        context["sum_total_paradas"] = sum_total_paradas
        context["sum_total_paradas_ativas"] = sum_total_paradas_ativas
        context["total_veiculos_cadastrados"] = total_veiculos_cadastrados
        context["sum_total_notificacoes"] = sum_total_notificacoes
        context["sum_total_notificacoes_lidas"] = sum_total_notificacoes_lidas
        context["sum_total_infracoes"] = sum_total_infracoes
        context["taxa_notificacoes"] = format(taxa_notificacoes, '.2f')
        context["taxa_notificacoes_lidas"] = format(taxa_notificacoes_lidas, '.2f')
        context["taxa_infracoes"] = format(taxa_infracoes, '.2f')
        context["horas_validas"] = horas_validas
        context["total_horas_paradas"] = total_horas_paradas
        context["delta_horas_paradas"] = delta_horas_paradas
        context["pesquisasFuncionarios"] = pesquisasFuncionarios
        context["vendasFuncionarios"] = vendasFuncionarios
        return render(request, self.template_name, context=context)
    
    def post(self, request):
        return render(request, self.template_name)


class Historico(View):
    template_name = 'dashboard-historico.html'
    
    def get(self, request):
        context = {}
        if 'limpar' in request.GET:
            ano_inicio = datetime.datetime.now().year
            mes_inicio = datetime.datetime.now().month
        else:    
            ano_inicio = request.GET.get('ano', None)
            mes_inicio = request.GET.get('mes', None)
            if not ano_inicio: ano_inicio = datetime.datetime.now().year
            if not mes_inicio: mes_inicio = datetime.datetime.now().month
        ano_fim = int(ano_inicio) - 1
        mes_fim = mes_inicio
        
        # NOTIFICAÇÕES E INFRAÇÕES
        tp_notificacao = TipoNotificacao.objects.filter(id__in=[1,2])
        tp_infracao = TipoNotificacao.objects.filter(id__in=[3,4])

        # PARADAS
        paradas = Parada.objects.filter(data_parada__year__gte=ano_fim, data_parada__month__gte=mes_fim, data_parada__year__lte=ano_inicio, data_parada__month__lte=mes_inicio)
        paradas = paradas.annotate(ano=TruncYear('data_parada'), mes=TruncMonth('data_parada'))
        paradas = paradas.values('ano', 'mes').annotate(total=Count('mes'))
        
        # NOTIFICACOES
        notificacoes = Notificacao.objects.filter(tipo_notificacao__in=tp_notificacao,data_notificacao__year__gte=ano_fim, data_notificacao__month__gte=mes_fim, data_notificacao__year__lte=ano_inicio, data_notificacao__month__lte=mes_inicio)
        notificacoes = notificacoes.annotate(ano=TruncYear('data_notificacao'), mes=TruncMonth('data_notificacao'))
        notificacoes = notificacoes.values('ano', 'mes').annotate(total=Count('mes'))

        # INFRACOES
        infracoes = Notificacao.objects.filter(tipo_notificacao__in=tp_infracao,data_notificacao__year__gte=ano_fim, data_notificacao__month__gte=mes_fim, data_notificacao__year__lte=ano_inicio, data_notificacao__month__lte=mes_inicio)
        infracoes = infracoes.annotate(ano=TruncYear('data_notificacao'), mes=TruncMonth('data_notificacao'))
        infracoes = infracoes.values('ano', 'mes').annotate(total=Count('mes'))

        # RETORNO
        context["paradas"] = paradas
        context["notificacoes"] = notificacoes
        context["infracoes"] = infracoes
        return render(request, self.template_name, context=context)


class Report(View):
    template_name = 'report.html'
    
    def get(self, request):
        if 'limpar' in request.GET:
            ano = None
            mes = None
        else:    
            ano = request.GET.get('ano', None)
            mes = request.GET.get('mes', None)

        paradas = Parada.objects.all()
        if ano:
            paradas = paradas.filter(data_parada__year=ano)
        if mes:
            paradas = paradas.filter(data_parada__month=mes)
        return render(request, self.template_name, { "paradas": paradas, })
    
    
    def post(self, request):
        ano = request.POST.get('ano', None)
        mes = request.POST.get('mes', None)

        paradas = Parada.objects.all()
        if ano:
            paradas = paradas.filter(data_parada__year=ano)
            
        if mes:
            paradas = paradas.filter(data_parada__month=mes)
        
        kwargs = {"paradas": paradas }
        return ExportarCSV(request, **kwargs)


def ExportarCSV(self, **kwargs):

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename={date}-paradas.xlsx'.format(date=datetime.datetime.now().strftime('%Y-%m-%d'),)
    response['paradas'] = kwargs['paradas']
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Paradas'

    # DEFINE OS CABECALHOS
    columns = ['Placa','Usuario','Data','Hora','HoraEstacionado','Valor']
    row_num = 1

    # INSERE OS CABECALHOS
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for parada in kwargs['paradas']:

        row_num += 1
        if parada.user:
            row = [parada.veiculo.placa,parada.user.first_name,parada.data_parada,parada.hora_parada,parada.quantidade_horas.descricao_horas,parada.quantidade_horas.valor]
        else:
            row = [parada.veiculo.placa,None,parada.data_parada,parada.hora_parada,parada.quantidade_horas.descricao_horas,parada.quantidade_horas.valor]
        
        # INCLUI OS DADOS NO ARQUIVO
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def keepalive(request):
    html = "<html><body>KeepAlive Ok.</body></html>"
    return HttpResponse(html)